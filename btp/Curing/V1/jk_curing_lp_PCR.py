"""
JK Tyre BTP — PCR Curing LP Scheduler v2
Designed By — Paranjay Dodiya — Algo8 AI Pvt Ltd
===============================================
===============================================
Full production-ready LP-based monthly curing schedule generator.

Architecture
------------
Phase 1 : ETL          — load & clean all input data
Phase 2 : LP Solve     — globally optimal press-minute allocation
Phase 3 : Rounding     — convert continuous LP solution to integer cycles
Phase 4 : Schedule     — build shift-wise row-level schedule
Phase 5 : Export       — Excel output (Dashboard / Machine Schedule /
                          Machine Utilization / Demand Fulfillment)

Key design decisions
--------------------
* LP objective  : minimise total unmet demand-minutes (sum of slack vars) 
                  while minimising the changeovers and Maximise the demand fulfillment
* Constraints   : machine capacity, SKU-machine eligibility (bounds),
                  demand coverage with slack
* Rounding      : floor to complete cycles only (no partial cycles)
* Priorities    : handled post-LP via tiebreaker in rounding — high-priority
                  SKUs get first pick of residual capacity after LP solve
* Cavities      : configurable per SKU (defaults to 2 for PCR)
* Shifts        : A (07–15h), B (15–23h), C (23–07h)


Addresses all 6 production constraints:

  1. Changeover time  — 300 min deducted from effective press capacity
                        per SKU switch on a machine
  2. Mould cleaning   — 120 min block inserted after every 6000 units
                        (3000 cycles × 2 moulds per press)
  3. Min changeovers  — small penalty per active (machine, SKU) pair
                        in LP objective concentrates SKUs on fewer machines
  4. Press shifting   — LP globally reallocates; continuity block locks
                        currently running moulds for their remaining life
  5. Max CO/shift     — configurable cap on changeovers per shift
                        (default 4); extras queued to next shift
  6. Mould tracking   — MouldTracker validates every assignment against
                        mould-SKU compatibility and availability
"""

import ast
import math
import warnings
from collections import defaultdict
from datetime import datetime, timedelta

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from scipy.optimize import linprog

try:
    from sqlalchemy import create_engine
except ImportError:
    create_engine = None

warnings.filterwarnings("ignore")


# ══════════════════════════════════════════════════════════════════════════════
# CONFIGURATION
# ══════════════════════════════════════════════════════════════════════════════
class Config:
    # ── database ──────────────────────────────────────────────────────────────
    DB_SERVER   = "35.208.174.2"
    DB_NAME     = "jkplanning_CTP"
    DB_USER     = "root"
    DB_PASSWORD = "Dev112233"

    # ── planning horizon ──────────────────────────────────────────────────────
    PLANNING_DAYS   = 28
    SHIFTS_PER_DAY  = 3
    HOURS_PER_SHIFT = 8
    SHIFT_START_HOUR = 7          # Shift A starts at 07:00

    # ── press & mould ─────────────────────────────────────────────────────────
    CAVITIES_PER_MOULD  = 2       # tyres produced per cure cycle
    MOULDS_PER_PRESS    = 2       # LH + RH mould per press
    NEW_MOULD_LIFE      = 3000    # cycles before mould needs cleaning

    # ── downtime constants (minutes) ──────────────────────────────────────────
    CHANGEOVER_DURATION_MIN = 360  # 420 min changeover + 60 min FTC check
    CLEANING_DURATION_MIN   = 180  # mould cleaning duration
    LOAD_UNLOAD_BUFFER_MIN  = 2.3  # added to raw cure time
    PRESS_EFFICIENCY        = 0.90  # divisor for warmup/transition losses

    # ── changeover scheduling ─────────────────────────────────────────────────
    MAX_CHANGEOVERS_PER_SHIFT = 3    # plant-level cap — configurable
    CHANGEOVER_PENALTY_WEIGHT = 0.01 # LP objective penalty per active (m, s) pair
                                      # relative to demand-minutes scale
    PLAN_DATE = datetime(2026, 2, 1, 7, 0, 0)
    # ── output ────────────────────────────────────────────────────────────────
    TYRE_TYPE   = "pcr"
    OUTPUT_FILE = f"CTP_PCR_Curing_LP_v4_PlanSchedule_Feb_{PLAN_DATE.date()}_28Days.xlsx"

    @classmethod
    def avail_mins(cls) -> float:
        return cls.PLANNING_DAYS * cls.SHIFTS_PER_DAY * cls.HOURS_PER_SHIFT * 60

    @classmethod
    def units_per_cleaning_cycle(cls) -> int:
        """Units produced before a full press cleaning is needed."""
        return cls.NEW_MOULD_LIFE * cls.CAVITIES_PER_MOULD * cls.MOULDS_PER_PRESS


# ══════════════════════════════════════════════════════════════════════════════
# MOULD TRACKER
# ══════════════════════════════════════════════════════════════════════════════
class MouldTracker:
    """
    Tracks mould availability, compatibility and life.

    Each mould has:
        compatible_skus  : set of SKU codes this mould can produce
        life_remaining   : cycles remaining before cleaning (default 3000)
        assigned_machine : machine currently using this mould (None = free)

    Rules enforced:
        - A (machine, SKU) assignment is only valid if ≥ MOULDS_PER_PRESS
          compatible, unassigned moulds exist for that SKU.
        - When a mould is assigned, it is locked to that machine.
        - After 3000 cycles the mould needs cleaning (120 min) — tracked in
          ScheduleBuilder.
    """

    def __init__(self):
        # mould_id -> dict
        self._ledger: dict[str, dict] = {}

    def load_from_df(self, df_mould: pd.DataFrame, df_running: pd.DataFrame):
        """
        Initialise ledger from:
          df_mould   : Master_Mapping_Mould_SKU (MouldNo, Matl.Code, Active Flag)
          df_running : currently running moulds (Machine, SKUCode, MouldNo, MouldLife_remaining)
        """
        # Build base ledger from mould master
        mould_col = "MouldNo" if "MouldNo" in df_mould.columns else "Mould"
        for _, row in df_mould.iterrows():
            mid = str(row[mould_col])
            sku = str(row["Matl.Code"])
            if mid not in self._ledger:
                self._ledger[mid] = {
                    "compatible_skus":   set(),
                    "life_remaining":    Config.NEW_MOULD_LIFE,
                    "assigned_machine":  None,
                }
            self._ledger[mid]["compatible_skus"].add(sku)

        # Update life and assignment from running moulds
        # df_running now has MouldNos as a list (both LH and RH moulds per machine)
        for _, row in df_running.iterrows():
            machine = str(row["Machine"])
            life    = int(row.get("MouldLife_remaining", Config.NEW_MOULD_LIFE))

            # Support both old single-mould (MouldNo) and new list (MouldNos) format
            raw = row.get("MouldNos", row.get("MouldNo", None))
            if raw is None:
                continue
            moulds = raw if isinstance(raw, list) else [str(raw)]

            for mould in moulds:
                mould = str(mould).strip()
                if mould in self._ledger:
                    self._ledger[mould]["life_remaining"]   = life
                    self._ledger[mould]["assigned_machine"] = machine

    def load_from_excel(self, mould_path: str, running_path: str = None):
        df_mould = pd.read_excel(mould_path)
        df_mould = df_mould[df_mould.get("Active Flag", pd.Series([True]*len(df_mould))).astype(bool)]
        df_mould = df_mould.rename(columns={"Matl.Code": "Matl.Code"})
        # Build simplified running df if path given
        if running_path:
            df_running = pd.read_excel(running_path)
            if "MouldLife_remaining" not in df_running.columns:
                df_running["MouldLife_remaining"] = Config.NEW_MOULD_LIFE
        else:
            df_running = pd.DataFrame(columns=["Machine", "SKUCode", "MouldNo", "MouldLife_remaining"])
        self.load_from_df(df_mould, df_running)

    def available_moulds_for_sku(self, sku: str) -> list[str]:
        """Return list of free mould IDs compatible with this SKU."""
        return [
            mid for mid, data in self._ledger.items()
            if sku in data["compatible_skus"]
            and data["assigned_machine"] is None
        ]

    def can_assign(self, sku: str) -> bool:
        """True if ≥ MOULDS_PER_PRESS free compatible moulds exist."""
        return len(self.available_moulds_for_sku(sku)) >= Config.MOULDS_PER_PRESS

    def get_eligible_machines_with_moulds(
        self, sku: str, candidate_machines: list[str]
    ) -> list[str]:
        """
        Filter candidate machines to those where a valid mould assignment
        is possible — i.e. free moulds exist for the SKU.
        (Machine-level physical compatibility is already in allowable matrix;
        this layer adds mould pool feasibility.)
        """
        if not self.can_assign(sku):
            return []
        return candidate_machines

    def assign_moulds(self, sku: str, machine: str) -> list[str]:
        """
        Assign MOULDS_PER_PRESS moulds to a machine for a SKU.
        Returns the list of assigned mould IDs.
        Raises ValueError if insufficient moulds available.
        """
        avail = self.available_moulds_for_sku(sku)
        if len(avail) < Config.MOULDS_PER_PRESS:
            raise ValueError(
                f"Cannot assign {Config.MOULDS_PER_PRESS} moulds for "
                f"SKU={sku} on Machine={machine}: only {len(avail)} free"
            )
        chosen = sorted(avail, key=lambda m: -self._ledger[m]["life_remaining"])[: Config.MOULDS_PER_PRESS]
        for mid in chosen:
            self._ledger[mid]["assigned_machine"] = machine
        return chosen

    def release_moulds(self, mould_ids: list[str]):
        """Free moulds at end of run so they can be re-assigned."""
        for mid in mould_ids:
            if mid in self._ledger:
                self._ledger[mid]["assigned_machine"] = None

    def mould_life(self, mould_id: str) -> int:
        return self._ledger.get(mould_id, {}).get("life_remaining", Config.NEW_MOULD_LIFE)

    def avg_life_remaining_for_sku(self, sku: str) -> float:
        moulds = self.available_moulds_for_sku(sku)
        if not moulds:
            return 0.0
        return sum(self._ledger[m]["life_remaining"] for m in moulds) / len(moulds)

    @property
    def summary(self) -> pd.DataFrame:
        rows = []
        for mid, d in self._ledger.items():
            rows.append({
                "MouldNo":          mid,
                "Compatible_SKUs":  ", ".join(sorted(d["compatible_skus"])),
                "Life_Remaining":   d["life_remaining"],
                "Assigned_Machine": d["assigned_machine"] or "FREE",
            })
        return pd.DataFrame(rows)


# ══════════════════════════════════════════════════════════════════════════════
# ETL
# ══════════════════════════════════════════════════════════════════════════════
class ETL:
    def __init__(self, engine=None, tyre_type: str = "pcr"):
        self.engine = engine
        self.t = tyre_type

    def _sql(self, q): return pd.read_sql(q, self.engine)

    # ── from DB ───────────────────────────────────────────────────────────────
    def load_demand(self, csv_path: str) -> pd.DataFrame:
        df = pd.read_csv(csv_path)
        df = (df.groupby("SKUCode")
                .agg(Quantity=("Updated_Requirement","sum"),
                     Priority=("ConsolidatedPriorityScore","max"))
                .reset_index())

        df = df[df["Quantity"] > 0].copy()
        df.to_excel("load_demand.xlsx", index=False)
        return df

    def load_cycle_times(self) -> pd.DataFrame:
        df = self._sql(
            f"SELECT Sapcode AS SKUCode, `Cure Time` AS Raw "
            f"FROM {Config.DB_NAME}.Master_Curing_Design_CycleTime_pcr"
        )
        df["CycleTime_min"] = np.round(
            (df["Raw"] + Config.LOAD_UNLOAD_BUFFER_MIN) / Config.PRESS_EFFICIENCY
        )
        df = df[["SKUCode","CycleTime_min"]].drop_duplicates("SKUCode")
        df['SKUCode'] = df['SKUCode'].str.strip()

        df.to_excel("load_cycle_times.xlsx", index=False)
        return df

    def load_machine_allowable(self) -> pd.DataFrame:
        df = self._sql(
            f"SELECT * FROM {Config.DB_NAME}"
            f".Master_Curing_Allowable_Machines_source_pcr"
        )
        df = df.rename(columns={"SKUCode":"SKUCode"})
        mcols = [c for c in df.columns if str(c).isdigit()]
        df["Machines"] = df.apply(
            lambda r: [str(c) for c in mcols if str(r[c]).strip().lower()=="yes"], axis=1
        )
        df = df[["SKUCode","Machines"]]
        #print(df.values)
        #df['Machines'] = df['Machines'].apply(lambda x: [int(i) for i in ast.literal_eval(x)])
        df['Machines'] = df['Machines'].apply(lambda lst: list(map(int, lst)))
              
        df.to_excel("load_machine_allowable.xlsx", index=False)
        return df

    def load_gt_inventory(self) -> pd.DataFrame:
        df = self._sql(
            f"SELECT ItemCode AS SKUCode, TotalQuantity AS GT_Inventory"
            f" FROM {Config.DB_NAME}.gt_inventory_manual_pcr"
        )

        df.to_excel("load_gt_inventory.xlsx", index=False)
        return df

    # def load_running_moulds(self) -> pd.DataFrame:
    #     """
    #     Returns one row per machine with columns:
    #         Machine | SKUCode | MouldNos (list) | MouldLife_remaining | Num_Moulds

    #     Each physical press has a LH and RH mould (WCNAME has LH/RH suffix).
    #     We strip the suffix to get the base machine ID, then group BOTH moulds
    #     together into a list.  Machines with only one mould (one side only) are
    #     handled naturally — MouldNos will be a single-element list.

    #     MouldLife_remaining = min(LH_life, RH_life): the press needs cleaning
    #     as soon as EITHER mould expires, so we track the worst-case life.
    #     """
        
    #     wc_master = self._sql(f"SELECT * FROM {Config.DB_NAME}.Master_WC_Master")
        
    #     wc_master = wc_master[['wcID', 'WCNAME']]

    #     df = self._sql(f"SELECT * FROM {Config.DB_NAME}.Daily_Running_Moulds_pcr")
    #     # df = df.drop(columns=["updatedAt"])

    #     dff = df[['WCNAME', 'Side','Sapcode', 'Current MouldNo', 'Mould life']]
    #     # dff['Mould life'] = 3000 - dff['Mould life']
    #     dff['Mould life'] = np.where(dff['Mould life']<0, 0, dff['Mould life']) 

    #     dff = dff.merge(wc_master, on=['WCNAME'], how='left')
    #     dff['WCNAME'] = dff['WCNAME'].str.replace(r'(LH|RH)$', '', regex=True).str.strip()
    #     dff['curing_machine'] = dff['WCNAME'] + dff['Side']

    #     Running_number_molds_sku = dff.groupby(['Sapcode']).agg({'Current MouldNo':'count'}).reset_index()
    #     Running_number_molds_sku.rename(columns={'Current MouldNo':'Noof_moulds'}, inplace=True)

    #     Running_Moulds = dff[['curing_machine', 'Current MouldNo', 'Sapcode', 'Mould life']]
    #     Running_Moulds.columns = ['WCNAME', 'Current MouldNo', 'Sapcode', 'Mould life']

    #     Running_Moulds['WCNAME'] = Running_Moulds['WCNAME'].str.strip('LH|RH')

    #     Running_Moulds['No'] = 1

    #     grouped = (
    #                 Running_Moulds.groupby("WCNAME")
    #                     .agg(
    #                         SKUCode=("Sapcode", "first"),
    #                         MouldNos=("Current MouldNo", list),
    #                         MouldLife_remaining=("Mould life", "min"), 
    #                         Num_Moulds=("No", "count"),
    #                     )
    #                     .reset_index()
    #             )

    #     grouped.columns = ["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"]
    #     #grouped['MouldNos'] = grouped['MouldNos'].apply(ast.literal_eval)

    #     grouped.to_excel("load_running_moulds.xlsx", index=False)
    #     return grouped[["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"]]

    def load_running_moulds(self) -> pd.DataFrame:
        """
        Returns one row per machine with columns:
            Machine | SKUCode | MouldNos (list) | MouldLife_remaining | Num_Moulds

        Updated:
        - Handles new DB format (no Side column)
        - Splits moulds using '#'
        - Recreates Side column (LH/RH) internally
        """

        wc_master = self._sql(f"SELECT * FROM {Config.DB_NAME}.Master_WC_Master")
        wc_master = wc_master[['wcID', 'WCNAME']]

        df = self._sql(f"SELECT * FROM {Config.DB_NAME}.Daily_Running_Moulds_pcr")

        # OLD LOGIC (Side-based) — NOT USED NOW
        # dff = df[['WCNAME', 'Side','Sapcode', 'Current MouldNo', 'Mould life']]

        # NEW LOGIC
        dff = df[['WCNAME', 'Sapcode', 'Current MouldNo', 'Mould life']].copy()

        dff['Mould life'] = np.where(dff['Mould life'] < 0, 0, dff['Mould life'])

        # datatype fix
        dff['WCNAME'] = dff['WCNAME'].astype(str)
        wc_master['WCNAME'] = wc_master['WCNAME'].astype(str)

        dff = dff.merge(wc_master, on=['WCNAME'], how='left')

        dff['WCNAME'] = dff['WCNAME'].str.replace(r'(LH|RH)$', '', regex=True).str.strip()

        # ❌ OLD LOGIC
        # dff['curing_machine'] = dff['WCNAME'] + dff['Side']

        # ✅ NEW LOGIC
        dff['curing_machine'] = dff['WCNAME']

        Running_number_molds_sku = dff.groupby(['Sapcode']).agg({'Current MouldNo': 'count'}).reset_index()
        Running_number_molds_sku.rename(columns={'Current MouldNo': 'Noof_moulds'}, inplace=True)

        Running_Moulds = dff[['curing_machine', 'Current MouldNo', 'Sapcode', 'Mould life']]
        Running_Moulds.columns = ['WCNAME', 'Current MouldNo', 'Sapcode', 'Mould life']

        # =============================================================================
        # ✅ NEW LOGIC: Split moulds + create Side column
        # =============================================================================
        split_df = Running_Moulds.copy()

        split_cols = split_df['Current MouldNo'].str.split('#', n=1, expand=True)

        split_df['Mould1'] = split_cols[0]
        split_df['Mould2'] = split_cols[1] if split_cols.shape[1] > 1 else None

        # LH rows
        df1 = split_df.copy()
        df1['Current MouldNo'] = df1['Mould1']
        df1['Side'] = 'LH'

        # RH rows
        df2 = split_df.copy()
        df2['Current MouldNo'] = df2['Mould2']
        df2['Side'] = 'RH'

        # Combine
        Running_Moulds = pd.concat([df1, df2], ignore_index=True)

        # Remove null moulds (important)
        Running_Moulds = Running_Moulds[Running_Moulds['Current MouldNo'].notna()]

        # Keep columns (Side is kept internally if needed later)
        Running_Moulds = Running_Moulds[['WCNAME', 'Current MouldNo', 'Side', 'Sapcode', 'Mould life']]

        # =============================================================================
        # SAME OLD AGGREGATION (UNCHANGED)
        # =============================================================================
        Running_Moulds['No'] = 1

        grouped = (
            Running_Moulds.groupby("WCNAME")
            .agg(
                SKUCode=("Sapcode", "first"),
                MouldNos=("Current MouldNo", list),
                MouldLife_remaining=("Mould life", "min"),
                Num_Moulds=("No", "count"),
            )
            .reset_index()
        )

        grouped.columns = ["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"]

        grouped.to_excel("load_running_moulds.xlsx", index=False)

        return grouped[["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"]]

    def load_mould_master(self) -> pd.DataFrame:
        df = self._sql(
            f"SELECT * FROM {Config.DB_NAME}.Master_Mapping_Mould_SKU "
            "WHERE `Active Flag`=True"
        )

        df.to_excel("load_mould_master.xlsx", index=False)
        return df

    # ── from Excel (offline / testing) ────────────────────────────────────────
    @staticmethod
    def load_demand_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        df = df.rename(columns={"Penetration%":"Priority"})
        return df[df["Quantity"] > 0].copy()

    @staticmethod
    def load_cycle_times_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        df = df.rename(columns={"Sapcode":"SKUCode","Cure Time":"Raw"})
        df["CycleTime_min"] = np.round(
            (df["Raw"] + Config.LOAD_UNLOAD_BUFFER_MIN) / Config.PRESS_EFFICIENCY
        )
        return df[["SKUCode","CycleTime_min"]].drop_duplicates("SKUCode")

    @staticmethod
    def load_machine_allowable_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        df["Machines"] = df["Machines"].apply(
            lambda x: ast.literal_eval(x) if isinstance(x, str) else []
        )
        return df[["SKUCode","Machines"]]

    @staticmethod
    def load_gt_inventory_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        return df.rename(columns={"Sapcode":"SKUCode","ctb_qty":"GT_Inventory"})[
            ["SKUCode","GT_Inventory"]
        ]

    @staticmethod
    def load_mould_master_from_excel(path: str) -> pd.DataFrame:
        df = pd.read_excel(path)
        if "Active Flag" in df.columns:
            df = df[df["Active Flag"].astype(bool)]
        return df

    @staticmethod
    def load_running_moulds_from_excel(path: str) -> pd.DataFrame:
        """
        Reads running moulds from Excel.  The source Excel may have one row
        per WCNAME (i.e. 4501LH and 4501RH as separate rows) matching the DB
        format, or may already be pre-aggregated.  We normalise to one row
        per Machine with MouldNos as a list, same as load_running_moulds().
        """
        df = pd.read_excel(path)
        if "MouldLife_remaining" not in df.columns:
            df["MouldLife_remaining"] = Config.NEW_MOULD_LIFE

        # Detect if WCNAME-style (LH/RH suffixed) or already machine-level
        wcname_col = next((c for c in ["WCNAME", "Machine"] if c in df.columns), None)
        mould_col  = next((c for c in ["Current MouldNo", "MouldNo", "MouldNos"] if c in df.columns), None)
        sku_col    = next((c for c in ["Sapcode", "SKUCode"] if c in df.columns), None)

        if wcname_col is None or mould_col is None or sku_col is None:
            raise ValueError(
                f"Running moulds Excel missing required columns. "
                f"Found: {list(df.columns)}"
            )

        df["Machine"] = (df[wcname_col].astype(str)
                         .str.replace(r"(LH|RH)$", "", regex=True).str.strip())
        df = df.rename(columns={sku_col: "SKUCode", mould_col: "MouldNo"})

        # If MouldNos is already a list column, return as-is
        if df["MouldNo"].dtype == object and isinstance(df["MouldNo"].iloc[0], list):
            return df[["Machine", "SKUCode", "MouldNos", "MouldLife_remaining"]]

        grouped = (
            df.groupby("Machine")
              .agg(
                  SKUCode=("SKUCode", "first"),
                  MouldNos=("MouldNo", list),
                  MouldLife_remaining=("MouldLife_remaining", "min"),
                  Num_Moulds=("MouldNo", "count"),
              )
              .reset_index()
        )
        single = (grouped["Num_Moulds"] == 1).sum()
        double = (grouped["Num_Moulds"] == 2).sum()
        print(f"  [ETL] Running moulds (Excel): {len(grouped)} machines | "
              f"2-mould: {double} | 1-mould: {single}")
        return grouped[["Machine", "SKUCode", "MouldNos", "MouldLife_remaining", "Num_Moulds"]]


# ══════════════════════════════════════════════════════════════════════════════
# LP SOLVER  (v2 — with changeover penalty + effective capacity)
# ══════════════════════════════════════════════════════════════════════════════
class LP_Solver:
    """
    Variables
    ---------
    x[s,m]   : press-minutes assigned to SKU s on machine m
    slack[s] : unmet demand-minutes for SKU s

    Objective
    ---------
    minimise  Σ slack[s]
            + CHANGEOVER_PENALTY_WEIGHT * Σ (x[s,m] / demand_mins[s])

    The second term penalises spreading SKU s across many machines
    (each active pair gets a fractional penalty proportional to how much
    of the demand it handles), which drives the LP to concentrate each
    SKU on as few machines as possible — i.e. fewer changeovers.

    Effective capacity
    ------------------
    Each machine has theoretical capacity = AVAIL_MINS.
    We reserve CHANGEOVER_DURATION_MIN for each *additional* SKU assigned
    beyond the first.  Because the number of SKUs per machine is unknown
    before solving, we use a conservative estimate:

        effective_cap[m] = AVAIL_MINS
                         - CHANGEOVER_DURATION_MIN * (expected_skus_per_m - 1)

    expected_skus_per_m is initialised to 1 and iterated once if needed.
    In practice, with the penalty in the objective the LP naturally limits
    SKUs per machine, so one iteration suffices.
    """

    def __init__(self):
        self.avail_mins = Config.avail_mins()
        self.co_mins    = Config.CHANGEOVER_DURATION_MIN
        self.penalty    = Config.CHANGEOVER_PENALTY_WEIGHT

    def solve(
        self,
        df_valid: pd.DataFrame,
        all_machines: list[str],
        mould_tracker: MouldTracker,
        locked_machine_mins: dict[str, float] = None,
    ) -> tuple[np.ndarray, dict]:
        """
        Parameters
        ----------
        df_valid            : SKU rows with Demand_Mins, Machines, CycleTime_min
        all_machines        : sorted list of all machine IDs
        mould_tracker       : for mould-level eligibility filtering
        locked_machine_mins : {machine: minutes} already consumed by
                              continuity blocks (currently running moulds)
        """
        locked = locked_machine_mins or {}
        S = len(df_valid)
        M = len(all_machines)
        midx = {m: i for i, m in enumerate(all_machines)}
        sku_rows = list(df_valid.itertuples(index=False))

        def xidx(s, mi): return s * M + mi

        # Effective capacity = avail - locked - conservatively reserve 1 CO
        # (LP penalty keeps actual SKUs/machine low, so 1 CO buffer is enough)
        eff_cap = {}
        for m in all_machines:
            cap = self.avail_mins - locked.get(str(m), 0.0) #- self.co_mins
            eff_cap[m] = max(cap, 0.0)

        n_vars = S * M + S
        # ── objective ────────────────────────────────────────────────────────
        c = np.zeros(n_vars)
        for s in range(S):
            c[S * M + s] = 1.0          # slack penalty (hard)
        for s, row in enumerate(sku_rows):
            if row.Demand_Mins > 0:
                for mi in range(M):
                    # soft penalty per active pair (encourages concentration)
                    c[xidx(s, mi)] += self.penalty / row.Demand_Mins

        # ── bounds ───────────────────────────────────────────────────────────
        bounds = [(0.0, None)] * n_vars
        for si, row in enumerate(sku_rows):
            sku = row.SKUCode
            eligible = set(row.Machines)
            # mould filter: remove machines if no compatible moulds available
            mould_eligible = set(
                mould_tracker.get_eligible_machines_with_moulds(sku, list(eligible))
            )
            for mi, mach in enumerate(all_machines):
                if mach not in mould_eligible:
                    bounds[xidx(si, mi)] = (0.0, 0.0)
                else:
                    bounds[xidx(si, mi)] = (0.0, float(row.Demand_Mins))

        # ── machine capacity  A_ub x ≤ b_ub ─────────────────────────────────
        A_cap = np.zeros((M, n_vars))
        b_cap = np.array([eff_cap[m] for m in all_machines])
        for mi in range(M):
            for si in range(S):
                A_cap[mi, xidx(si, mi)] = 1.0

        # ── demand coverage  -Σ x[s,m] - slack[s] ≤ -demand_mins[s] ─────────
        A_dem = np.zeros((S, n_vars))
        b_dem = np.zeros(S)
        for si, row in enumerate(sku_rows):
            for mi in range(M):
                A_dem[si, xidx(si, mi)] = -1.0
            A_dem[si, S * M + si] = -1.0
            b_dem[si] = -float(row.Demand_Mins)

        A_ub = np.vstack([A_cap, A_dem])
        b_ub = np.concatenate([b_cap, b_dem])

        print(f"  [LP] {n_vars:,} vars | {len(b_ub):,} constraints | "
              f"Penalty weight: {self.penalty}")
        print(f"  [LP] Eff capacity range: "
              f"{min(eff_cap.values()):,.0f}–{max(eff_cap.values()):,.0f} min/press")

        result = linprog(c, A_ub=A_ub, b_ub=b_ub, bounds=bounds, method="highs")
        if result.status != 0:
            raise RuntimeError(f"LP did not converge: {result.message}")

        unmet = sum(result.x[S * M + s] for s in range(S))
        print(f"  [LP] OPTIMAL | Unmet demand-mins: {unmet:,.0f} ({unmet/60:.1f} hrs)")

        meta = {"S":S,"M":M,"midx":midx,"all_machines":all_machines,
                "sku_rows":sku_rows,"xidx":xidx}
        return result.x, meta


# ══════════════════════════════════════════════════════════════════════════════
# ROUNDER  (v2 — changeover-aware)
# ══════════════════════════════════════════════════════════════════════════════
class Rounder:
    """
    1. Floor all LP allocations to integer cycles.
    2. Deduct actual changeover time from each machine that runs > 1 SKU.
    3. Top up under-fulfilled high-priority SKUs with residual capacity.
    4. Record (machine → list[SKU]) assignment order for ScheduleBuilder.
    """

    def __init__(self):
        self.avail_mins = Config.avail_mins()
        self.co_mins    = Config.CHANGEOVER_DURATION_MIN

    def round(
        self,
        x: np.ndarray,
        meta: dict,
        df_valid: pd.DataFrame,
        locked_machine_mins: dict[str, float],
    ) -> tuple[pd.DataFrame, dict[str, list[str]]]:
        """
        Returns
        -------
        df_sched         : (Machine, SKUCode, Priority, CycleTime_min,
                            Cycles, Units_Planned, Mins_Used, Days_Used)
        machine_sku_order: {machine: [sku1, sku2, ...]} in run order
        """
        S        = meta["S"]
        M        = meta["M"]
        machines = meta["all_machines"]
        xidx     = meta["xidx"]
        sku_rows = meta["sku_rows"]

        # Remaining press capacity after locked continuity blocks
        machine_cap = {
            m: self.avail_mins - locked_machine_mins.get(m, 0.0)
            for m in machines
        }

        # First pass: floor all allocations
        raw = []
        for si, row in enumerate(sku_rows):
            ct = row.CycleTime_min
            for mi, mach in enumerate(machines):
                mins_lp = x[xidx(si, mi)]
                if mins_lp < ct:
                    continue
                cycles     = int(mins_lp / ct)
                actual_min = cycles * ct
                raw.append({"si": si, "mi": mi, "mach": mach,
                             "sku": row.SKUCode, "ct": ct,
                             "cycles": cycles, "mins": actual_min,
                             "priority": row.Priority})

        # Deduct changeover time for machines with > 1 SKU
        # Group raw assignments by machine
        by_machine: dict[str, list] = defaultdict(list)
        for a in raw:
            by_machine[a["mach"]].append(a)

        # For each machine, sort SKUs by priority, compute CO cost, trim if needed
        final = []
        machine_sku_order: dict[str, list[str]] = {}

        for mach, assignments in by_machine.items():
            assignments.sort(key=lambda a: -a["priority"])
            n_skus   = len(assignments)
            co_cost  = max(0, n_skus - 1) * self.co_mins
            avail    = machine_cap[mach] - co_cost
            used     = 0.0

            mach_skus = []
            for a in assignments:
                if avail - used < a["ct"]:
                    continue
                max_mins  = avail - used
                max_cyc   = int(max_mins / a["ct"])
                cycles    = min(a["cycles"], max_cyc)
                if cycles <= 0:
                    continue
                actual_min = cycles * a["ct"]
                used      += actual_min
                mach_skus.append(a["sku"])
                final.append({
                    "Machine":       mach,
                    "SKUCode":       a["sku"],
                    "Priority":      round(a["priority"], 4),
                    "CycleTime_min": a["ct"],
                    "Cycles":        cycles,
                    "Units_Planned": cycles * Config.CAVITIES_PER_MOULD,
                    "Mins_Used":     round(actual_min, 1),
                    "Days_Used":     round(actual_min / (Config.SHIFTS_PER_DAY
                                          * Config.HOURS_PER_SHIFT * 60), 2),
                })
            machine_sku_order[mach] = mach_skus
            machine_cap[mach] -= used + co_cost

        # Second pass: top-up under-fulfilled high-priority SKUs
        planned = defaultdict(int)
        for a in final:
            planned[a["SKUCode"]] += a["Units_Planned"]

        for si, row in enumerate(sorted(
            enumerate(sku_rows), key=lambda x: -x[1].Priority
        )):
            si, row = si, row[1]
            sku    = row.SKUCode
            ct     = row.CycleTime_min
            needed = int(row.Demand) - planned[sku]
            if needed <= 0:
                continue
            eligible = sorted(
                [mi for mi, m in enumerate(machines)
                 if machine_cap[machines[mi]] >= ct
                 and m in set(row.Machines)],
                key=lambda mi: -machine_cap[machines[mi]],
            )
            for mi in eligible:
                if needed <= 0:
                    break
                mach = machines[mi]
                cap  = machine_cap[mach]
                # Adding a new SKU to this machine costs one changeover
                existing_skus = len(machine_sku_order.get(mach, []))
                co  = self.co_mins if existing_skus > 0 else 0.0
                available = cap - co
                if available < ct:
                    continue
                extra_c = min(int(available / ct), math.ceil(needed / Config.CAVITIES_PER_MOULD))
                if extra_c <= 0:
                    continue
                actual_min = extra_c * ct
                machine_cap[mach] -= actual_min + co
                units = extra_c * Config.CAVITIES_PER_MOULD
                planned[sku] += units
                needed       -= units
                machine_sku_order.setdefault(mach, []).append(sku)
                final.append({
                    "Machine":       mach,
                    "SKUCode":       sku,
                    "Priority":      round(row.Priority, 4),
                    "CycleTime_min": ct,
                    "Cycles":        extra_c,
                    "Units_Planned": units,
                    "Mins_Used":     round(actual_min, 1),
                    "Days_Used":     round(actual_min / (Config.SHIFTS_PER_DAY
                                          * Config.HOURS_PER_SHIFT * 60), 2),
                })

        df_sched = pd.DataFrame(final)
        total_co = sum(
            max(0, len(v) - 1) * self.co_mins
            for v in machine_sku_order.values()
        )
        print(f"  [Round] Rows: {len(df_sched)} | "
              f"Units: {df_sched['Units_Planned'].sum():,.0f} | "
              f"Total CO time: {total_co/60:.1f} hrs | "
              f"Changeovers: {sum(max(0,len(v)-1) for v in machine_sku_order.values())}")
        return df_sched, machine_sku_order


# ══════════════════════════════════════════════════════════════════════════════
# SCHEDULE BUILDER  (v2 — changeover + cleaning blocks + CO shift cap)
# ══════════════════════════════════════════════════════════════════════════════
class ScheduleBuilder:
    """
    Builds a shift-level timeline for each machine:

      PRODUCTION   : normal tyre production rows
      CHANGEOVER   : 300-min block between SKU runs
      MOULD_CLEAN  : 120-min block inserted after every
                     units_per_cleaning_cycle (default 6000) units

    Changeover cap: at most MAX_CHANGEOVERS_PER_SHIFT changeovers
    start in any single 8-hour shift across the whole plant.
    Excess changeovers are deferred to the next shift.
    """

    def __init__(self, plan_start: datetime):
        self.plan_start = plan_start
        self.plan_end   = plan_start + timedelta(days=Config.PLANNING_DAYS)
        self.max_co_shift = Config.MAX_CHANGEOVERS_PER_SHIFT
        # {(date, shift): count_of_COs_started}
        self._co_shift_counter: dict[tuple, int] = defaultdict(int)

    # ── shift helpers ─────────────────────────────────────────────────────────
    @staticmethod
    def _get_shift(dt: datetime) -> tuple[str, datetime]:
        return _get_shift_fn(dt)

    def _shift_key(self, dt: datetime) -> tuple:
        shift, _ = self._get_shift(dt)
        return (dt.date(), shift)

    def _next_co_slot(self, earliest: datetime) -> datetime:
        """
        Return the earliest datetime >= earliest at which a new changeover
        can start without exceeding MAX_CHANGEOVERS_PER_SHIFT.
        Defers to next shift if current shift is full.
        """
        dt = earliest
        for _ in range(Config.PLANNING_DAYS * Config.SHIFTS_PER_DAY + 1):
            key = self._shift_key(dt)
            if self._co_shift_counter[key] < self.max_co_shift:
                self._co_shift_counter[key] += 1
                return dt
            # Move to start of next shift
            _, shift_end = self._get_shift(dt)
            dt = shift_end
        return dt  # fallback — should not reach here

    # ── row maker ─────────────────────────────────────────────────────────────
    def _make_row(self, start: datetime, end: datetime, machine: str,
                  sku: str, qty: int, ct: float, remarks: str, gt_inv: int = 0):
        shift, _ = self._get_shift(start)
        return {
            "Date":          start.date(),
            "Shift":         shift,
            "Machine":       machine,
            "SKUCode":       sku,
            "StartTime":     start,
            "EndTime":       end,
            "Qty":           qty,
            "CycleTime_min": round(ct, 2),
            "GT_Inventory":  gt_inv,
            "Remarks":       remarks,
        }

    # ── production block splitter ─────────────────────────────────────────────
    def _split_block(
        self, start: datetime, end: datetime,
        machine: str, sku: str, ct: float,
        gt_inv: int, remarks: str, units_so_far: int,
    ) -> tuple[list[dict], int]:
        """
        Split a production block (start→end) into shift rows.
        Insert MOULD_CLEAN events every units_per_cleaning_cycle units.
        Returns (rows, new_units_so_far).
        """
        rows = []
        cleaning_cycle = Config.units_per_cleaning_cycle()
        total_mins = (end - start).total_seconds() / 60
        total_units = int(total_mins / ct) * Config.CAVITIES_PER_MOULD
        produced   = 0
        curr       = start

        while curr < end and produced < total_units:
            # How many units until next cleaning?
            units_to_clean = cleaning_cycle - (units_so_far % cleaning_cycle)
            units_this_run = min(total_units - produced, units_to_clean)
            mins_this_run  = math.ceil(units_this_run / Config.CAVITIES_PER_MOULD) * ct
            run_end        = min(curr + timedelta(minutes=mins_this_run), end)

            # Split run_end across shifts
            inner = curr
            run_produced = 0
            run_total = int((run_end - curr).total_seconds() / 60 / ct) * Config.CAVITIES_PER_MOULD

            while inner < run_end:
                _, shift_end = self._get_shift(inner)
                slice_end    = min(shift_end, run_end)
                dur          = (slice_end - inner).total_seconds() / 60
                if dur <= 0:
                    inner = slice_end
                    continue
                if slice_end == run_end:
                    qty = run_total - run_produced
                else:
                    qty = int(dur / ct) * Config.CAVITIES_PER_MOULD
                rows.append(self._make_row(inner, slice_end, machine, sku,
                                           qty, ct, remarks, gt_inv))
                run_produced += qty
                inner = slice_end

            produced      += run_produced
            units_so_far  += run_produced
            curr           = run_end

            # Insert cleaning if hit cycle boundary and more production follows
            if (units_so_far % cleaning_cycle == 0
                    and produced < total_units and curr < end):
                clean_end = curr + timedelta(minutes=Config.CLEANING_DURATION_MIN)
                rows.append(self._make_row(curr, clean_end, machine,
                                           "MOULD_CLEAN", 0, 0.0,
                                           f"Mould cleaning after {units_so_far} units"))
                curr = clean_end

        return rows, units_so_far

    # ── main builder ──────────────────────────────────────────────────────────
    def build(
        self,
        df_sched: pd.DataFrame,
        machine_sku_order: dict[str, list[str]],
        df_gt: pd.DataFrame,
        continuity_rows: list[dict],
    ) -> pd.DataFrame:
        """
        Parameters
        ----------
        df_sched          : rounded allocation (Machine, SKUCode, Cycles, ...)
        machine_sku_order : {machine: [sku1, sku2...]} run order
        df_gt             : GT inventory for each SKU
        continuity_rows   : pre-built rows for currently running moulds
        """
        gt_map      = dict(zip(df_gt["SKUCode"], df_gt["GT_Inventory"]))
        # Machine cursors: start at plan_start unless occupied by continuity
        machine_free: dict[str, datetime] = {}
        for r in continuity_rows:
            m = r["Machine"]
            if m not in machine_free or r["EndTime"] > machine_free[m]:
                machine_free[m] = r["EndTime"]

        # Build lookup: (machine, sku) -> Cycles
        alloc: dict[tuple, dict] = {}
        for _, row in df_sched.iterrows():
            alloc[(row["Machine"], row["SKUCode"])] = row.to_dict()

        #all_rows = list(continuity_rows)
        all_rows = []
        con_dataframe = con_split_into_shifts(pd.DataFrame(continuity_rows))

        con_dataframe['Date'] = con_dataframe['StartTime'].dt.date
        con_dataframe['Date'] = con_dataframe['Date'].astype('datetime64[ns]')
        con_dataframe['Date'] = np.where(con_dataframe['StartTime'].dt.hour.isin([0,1,2,3,4,5,6]), con_dataframe['Date'] - pd.Timedelta(1, inut='d'), con_dataframe['Date'])

        # Process each machine in order
        for mach, sku_order in machine_sku_order.items():
            cursor = machine_free.get(str(mach), self.plan_start)
            units_so_far = 0   # cumulative units for cleaning tracker

            for idx, sku in enumerate(sku_order):
                key = (mach, sku)
                if key not in alloc:
                    continue
                a  = alloc[key]
                ct = a["CycleTime_min"]
                gt = int(gt_map.get(sku, 0))

                # Changeover block (not before first SKU on this machine)
                if idx > 0:
                    co_start = self._next_co_slot(cursor)
                    co_end   = co_start + timedelta(minutes=Config.CHANGEOVER_DURATION_MIN)
                    all_rows.append(self._make_row(
                        co_start, co_end, mach, "CHANGEOVER", 0, 0.0,
                        f"C/O to {sku}"
                    ))
                    cursor = co_end

                # Production block
                total_mins = a["Mins_Used"]
                block_end  = min(
                    cursor + timedelta(minutes=total_mins),
                    self.plan_end,
                )
                if block_end <= cursor:
                    continue

                prod_rows, units_so_far = self._split_block(
                    cursor, block_end, mach, sku, ct,
                    gt, "LP Scheduled", units_so_far,
                )
                all_rows.extend(prod_rows)
                cursor = block_end
            machine_free[mach] = cursor

        df_out = pd.DataFrame(all_rows)
        df_out = pd.concat([df_out, con_dataframe], axis = 0)
        if not df_out.empty:
            df_out = df_out.sort_values(["Machine","StartTime"]).reset_index(drop=True)
        co_count = (df_out["SKUCode"] == "CHANGEOVER").sum()
        cl_count = (df_out["SKUCode"] == "MOULD_CLEAN").sum()
        prod_qty = df_out.loc[~df_out["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"]),"Qty"].sum()
        print(f"  [Build] Total rows: {len(df_out)} | "
              f"Prod qty: {prod_qty:,.0f} | "
              f"Changeovers: {co_count} | Cleanings: {cl_count}")
        return df_out


# ══════════════════════════════════════════════════════════════════════════════
# ORCHESTRATOR
# ══════════════════════════════════════════════════════════════════════════════
class JK_LP_Curing_Scheduler_v2:
    def __init__(self):
        self.avail_mins = Config.avail_mins()

    # ── prep ──────────────────────────────────────────────────────────────────
    def _prepare_skus(self, df_demand, df_cycles, df_allow, df_gt,
                      mould_tracker: MouldTracker, df_running):
        cycle_map  = dict(zip(df_cycles["SKUCode"], df_cycles["CycleTime_min"]))
        mach_map   = dict(zip(df_allow["SKUCode"],  df_allow["Machines"]))
        gt_map     = dict(zip(df_gt["SKUCode"],     df_gt["GT_Inventory"]))

        rows = []
        for _, r in df_demand.iterrows():
            sku      = r["SKUCode"]
            qty      = int(r["Quantity"])
            priority = r["Priority"]
            ct       = cycle_map.get(sku, 15)
            machines = mach_map.get(sku, [])
            gt       = int(gt_map.get(sku, 0))
            dm       = math.ceil(qty / Config.CAVITIES_PER_MOULD) * ct if ct else 0
            if sku in df_running['SKUCode'].unique():
                has_mould = True
            else:
                has_mould = mould_tracker.can_assign(sku) if (ct and machines) else False

            #has_mould = mould_tracker.can_assign(sku) if (ct and machines) else False
            schedulable = bool(ct and machines and has_mould)
            skip = ("" if schedulable
                    else ("No cycle time" if not ct
                          else ("No machine mapping" if not machines
                                else "No compatible mould available")))
            rows.append({
                "SKUCode":          sku,
                "Demand":           qty,
                "Priority":         priority,
                "GT_Inventory":     gt,
                "CycleTime_min":    ct,
                "Machines":         machines,
                "Num_Machines":     len(machines),
                "Demand_Mins":      dm,
                "Presses_Needed":   round(dm / self.avail_mins, 2) if ct else 0,
                "Schedulable":      schedulable,
                "Skip_Reason":      skip,
            })

        df_all   = pd.DataFrame(rows)
        df_valid = df_all[df_all["Schedulable"]].copy().reset_index(drop=True)
        df_valid = df_valid.sort_values(["Priority","Num_Machines"],
                                        ascending=[False,True]).reset_index(drop=True)
        all_machines = sorted({m for ml in df_allow["Machines"] for m in ml})
        print(f"  [Prep] Schedulable: {len(df_valid)}/{len(df_all)} | "
              f"Machines: {len(all_machines)} | "
              f"Demand: {df_valid['Demand'].sum():,.0f}")
        return df_valid, df_all, all_machines

    def _build_continuity(self, df_running: pd.DataFrame,
                          df_valid: pd.DataFrame,
                          df_gt: pd.DataFrame,
                          plan_start: datetime) -> tuple[list[dict], dict[str, float], dict[str, int]]:
        """
        Schedule continuity blocks for all currently running machines.

        Key rules
        ---------
        1. Group all machines by SKU — every press running SKU X works together
           to satisfy SKU X demand before the LP handles anything else.
        2. Distribute demand proportionally across machines, weighted by each
           machine's horizon capacity (more moulds / faster cycle = larger share).
        3. If total group capacity < demand (even full horizon), run all machines
           for the full horizon and return the unmet remainder to the LP.
        4. Mould cleaning (120 min) is inserted mid-run whenever a machine hits
           NEW_MOULD_LIFE cycles (min of LH/RH life drives the trigger).
        5. locked_machine_mins tells the LP how much capacity each machine has
           already committed.

        Returns
        -------
        continuity_rows    : list of shift-level row dicts
        locked_machine_mins: {machine: minutes_committed}
        demand_remainder   : {sku: units_still_needed_from_LP}
        """
        if df_running is None or df_running.empty:
            return [], {}, {}

        cycle_map  = dict(zip(df_valid["SKUCode"], df_valid["CycleTime_min"]))
        demand_map = dict(zip(df_valid["SKUCode"], df_valid["Demand"]))
        gt_map     = dict(zip(df_gt["SKUCode"],    df_gt["GT_Inventory"]))
        plan_end      = plan_start + timedelta(days=Config.PLANNING_DAYS)
        horizon_mins  = Config.avail_mins()

        continuity_rows  = []
        locked_mins      = {}
        demand_remainder = {}

        # ── Step 1: group running machines by SKU ─────────────────────────────
        sku_groups: dict[str, list] = defaultdict(list)
        for _, row in df_running.iterrows():
            mach = str(row["Machine"])
            sku  = str(row["SKUCode"])
            ct   = cycle_map.get(sku, 15)
            if not ct:
                locked_mins[mach] = 0.0   # unknown cycle time — release immediately
                continue
            sku_groups[sku].append({
                "machine":  mach,
                "life":     int(row.get("MouldLife_remaining", Config.NEW_MOULD_LIFE)),
                "cavities": int(row.get("Num_Moulds", Config.MOULDS_PER_PRESS)),
                "ct":       ct,
            })

        # ── Step 2: process each SKU group ────────────────────────────────────
        for sku, group in sku_groups.items():
            demand = demand_map.get(sku, 0)
            gt     = int(gt_map.get(sku, 0))

            if demand <= 0:
                for m in group:
                    locked_mins[m["machine"]] = 0.0
                continue

            # Total units each machine can produce over the full horizon
            for m in group:
                m["max_units"] = int(horizon_mins / m["ct"]) * m["cavities"]
            group_total_cap = sum(m["max_units"] for m in group)

            can_meet = group_total_cap >= demand

            # ── Step 3: distribute demand across machines ──────────────────────
            # Proportional by each machine's share of group capacity.
            # Process machines so rounding shortfall goes to the highest-cap press.
            group_sorted = sorted(group, key=lambda m: -m["max_units"])
            remaining_demand = demand
            for i, m in enumerate(group_sorted):
                if not can_meet:
                    # Run every machine at full horizon — LP handles remainder
                    alloc = m["max_units"]
                elif remaining_demand <= 0:
                    alloc = 0
                elif i == len(group_sorted) - 1:
                    # Last machine: take exactly what's left (no rounding gap)
                    alloc = min(remaining_demand, m["max_units"])
                else:
                    share = m["max_units"] / group_total_cap if group_total_cap else 0
                    alloc = min(math.ceil(demand * share), m["max_units"], remaining_demand)

                m["alloc_units"] = max(alloc, 0)
                remaining_demand -= m["alloc_units"]

            # Pass unmet demand to LP
            if not can_meet:
                unmet = demand - group_total_cap
                demand_remainder[sku] = max(int(unmet), 0)
                if unmet > 0:
                    print(f"  [Cont] SKU {sku}: group cap {group_total_cap:,} < "
                          f"demand {demand:,} -> {unmet:,} units to LP")
            else:
                demand_remainder[sku] = 0

            # ── Step 4: build shift rows for each machine ──────────────────────
            for m in group:
                mach        = m["machine"]
                ct          = m["ct"]
                life        = m["life"]
                cavities    = m["cavities"]
                alloc_units = m["alloc_units"]

                if alloc_units <= 0:
                    locked_mins[mach] = 0.0
                    continue

                alloc_mins         = math.ceil(alloc_units / cavities) * ct
                units_before_clean = life * cavities   # min(LH,RH) drives trigger
                cursor             = plan_start
                remaining_mins     = alloc_mins

                while remaining_mins > 0 and cursor < plan_end:
                    # Run until mould clean threshold or allocation exhausted
                    mins_to_clean = (units_before_clean / cavities) * ct
                    block_mins    = min(mins_to_clean, remaining_mins)
                    block_end     = min(cursor + timedelta(minutes=block_mins), plan_end)
                    actual_mins   = (block_end - cursor).total_seconds() / 60
                    qty           = int(actual_mins / ct) * cavities

                    shift, _ = _get_shift_fn(cursor)
                    continuity_rows.append({
                        "Date":          cursor.date(),
                        "Shift":         shift,
                        "Machine":       mach,
                        "SKUCode":       sku,
                        "StartTime":     cursor,
                        "EndTime":       block_end,
                        "Qty":           qty,
                        "CycleTime_min": ct,
                        "GT_Inventory":  gt,
                        "Remarks":       f"Continuity ({cavities}-mould press)",
                    })
                    cursor        = block_end
                    remaining_mins -= block_mins

                    # Mould cleaning if more production follows
                    if remaining_mins > 0 and cursor < plan_end:
                        clean_end = min(
                            cursor + timedelta(minutes=Config.CLEANING_DURATION_MIN),
                            plan_end,
                        )
                        shift, _ = _get_shift_fn(cursor)
                        continuity_rows.append({
                            "Date":          cursor.date(),
                            "Shift":         shift,
                            "Machine":       mach,
                            "SKUCode":       "MOULD_CLEAN",
                            "StartTime":     cursor,
                            "EndTime":       clean_end,
                            "Qty":           0,
                            "CycleTime_min": 0.0,
                            "GT_Inventory":  0,
                            "Remarks":       "Mould cleaning (continuity)",
                        })
                        cursor             = clean_end
                        units_before_clean = Config.NEW_MOULD_LIFE * cavities

                locked_mins[mach] = (cursor - plan_start).total_seconds() / 60

        demand_remainder = {k: (0 if v < 50 else v) for k, v in demand_remainder.items()}

        total_locked = sum(locked_mins.values())
        print(f"  [Cont] {len(continuity_rows)} rows | "
              f"Machines committed: {len(locked_mins)} | "
              f"Locked mins: {total_locked:,.0f} | "
              f"SKUs with LP remainder: {sum(1 for v in demand_remainder.values() if v > 0)}")
        return continuity_rows, locked_mins, demand_remainder


    def _build_summary(self, df_all, df_sched, df_shift):
        """
        Build demand fulfillment summary.

        Source of truth: df_shift (the shift schedule).
        It contains ALL production — both continuity rows and LP-sourced rows.
        We never mix df_mach (LP allocation) with df_shift here because
        ScheduleBuilder converts df_mach rows into df_shift rows, meaning
        every LP unit already appears in df_shift. Adding df_mach on top
        would double-count every unit.
        """
        if not df_shift.empty:
            prod = df_shift[~df_shift["SKUCode"].isin(["CHANGEOVER", "MOULD_CLEAN"])]
            planned = prod.groupby("SKUCode")["Qty"].sum().to_dict()
        else:
            planned = {}

        rows = []
        for _, r in df_all.iterrows():
            sku  = r["SKUCode"]
            d    = r["Demand"]
            plan = int(planned.get(sku, 0))
            gap  = max(d - plan, 0)
            pct  = round(plan / d * 100, 1) if d > 0 else 100.0
            if not r["Schedulable"]:
                status = "UNSCHEDULABLE"
            elif gap <= 0:
                status = "FULLY MET"
            elif plan > 0:
                status = "PARTIAL"
            else:
                status = "UNMET"
            rows.append({"SKUCode": sku, "Priority": r["Priority"],
                         "Demand": d, "GT_Inventory": r["GT_Inventory"],
                         "Planned_Units": plan, "Gap": gap,
                         "Fulfillment_Pct": pct, "Status": status,
                         "CycleTime_min": r["CycleTime_min"],
                         "Eligible_Machines": r["Num_Machines"],
                         "Presses_Needed": r["Presses_Needed"],
                         "Skip_Reason": r["Skip_Reason"]})
        return pd.DataFrame(rows).sort_values("Priority", ascending=False)

    def _build_util(self, df_sched, df_shift, all_machines):
        # Use df_shift as single source of truth (contains both continuity and LP rows).
        # Compute elapsed minutes from actual timestamps to avoid double-counting.
        if not df_shift.empty:
            prod = df_shift[~df_shift["SKUCode"].isin(["CHANGEOVER","MOULD_CLEAN"])].copy()
            prod['Machine'] = prod['Machine'].astype('int64')
            prod["Elapsed"] = (pd.to_datetime(prod["EndTime"]) - pd.to_datetime(prod["StartTime"])).dt.total_seconds() / 60
            grp = prod.groupby("Machine").agg(
                Used_Mins=("Elapsed",  "sum"),
                Total_Units=("Qty",    "sum"),
                SKUs_Count=("SKUCode", "nunique"),
            ).reset_index()
            # Cycles from LP allocation only (continuity doesn't track cycles)
            if not df_sched.empty:
                cyc = df_sched.groupby("Machine")["Cycles"].sum().reset_index()
                grp = grp.merge(cyc, on="Machine", how="left").fillna(0)
                grp.rename(columns={"Cycles":"Total_Cycles"}, inplace=True)
            else:
                grp["Total_Cycles"] = 0
        else:
            grp = pd.DataFrame(columns=["Machine","Used_Mins","Total_Units","SKUs_Count","Total_Cycles"])
        df_u = pd.DataFrame({"Machine": all_machines}).merge(grp, on="Machine", how="left").fillna(0)
        df_u["Available_Mins"]  = self.avail_mins
        df_u["Idle_Mins"]       = self.avail_mins - df_u["Used_Mins"]
        df_u["Utilization_Pct"] = ((df_u["Used_Mins"] / df_u['Available_Mins']) * 100).round(2)
        return df_u[["Machine","Available_Mins","Used_Mins","Idle_Mins",
                     "Utilization_Pct","SKUs_Count","Total_Cycles","Total_Units"]]               .sort_values("Utilization_Pct", ascending=False)

    def _print_results(self, df_summary, df_util, df_shift):
        td  = df_summary["Demand"].sum()
        tp  = df_summary["Planned_Units"].sum()
        co  = (df_shift["SKUCode"]=="CHANGEOVER").sum() if not df_shift.empty else 0
        cl  = (df_shift["SKUCode"]=="MOULD_CLEAN").sum() if not df_shift.empty else 0
        print(f"\n{'='*64}")
        print(f"  Total demand    : {td:>10,.0f}")
        print(f"  Units planned   : {tp:>10,.0f}  ({tp/td*100:.1f}%)")
        print(f"  Gap             : {td-tp:>10,.0f}")
        print(f"  Avg press util  : {df_util['Utilization_Pct'].mean()}%")
        print(f"  Changeover rows : {co:>10}")
        print(f"  Mould clean rows: {cl:>10}")
        print(f"  Fully met SKUs  : {(df_summary['Status']=='FULLY MET').sum():>10}")
        print(f"  Partial SKUs    : {(df_summary['Status']=='PARTIAL').sum():>10}")
        print(f"  Unmet SKUs      : {(df_summary['Status']=='UNMET').sum():>10}")
        print(f"  Unschedulable   : {(df_summary['Status']=='UNSCHEDULABLE').sum():>10}")
        print(f"{'='*64}")

    # ── main entry point ──────────────────────────────────────────────────────
    def run(
        self,
        df_demand:    pd.DataFrame,
        df_cycles:    pd.DataFrame,
        df_allow:     pd.DataFrame,
        df_gt:        pd.DataFrame,
        mould_tracker: MouldTracker,
        df_running:   pd.DataFrame = None,
        plan_start:   datetime     = None,
    ) -> dict:

        if plan_start is None:
            plan_start = datetime.now().replace(
                hour=Config.SHIFT_START_HOUR, minute=0, second=0, microsecond=0
            )

        print("\n" + "="*64)
        print(f"  JK Tyre PCR Curing LP Scheduler v2")
        print(f"  Plan start  : {plan_start:%Y-%m-%d %H:%M}")
        print(f"  Horizon     : {Config.PLANNING_DAYS} days ({self.avail_mins:,.0f} min/press)")
        print(f"  Changeover  : {Config.CHANGEOVER_DURATION_MIN} min | "
              f"Max/shift: {Config.MAX_CHANGEOVERS_PER_SHIFT}")
        print(f"  Mould clean : {Config.CLEANING_DURATION_MIN} min every "
              f"{Config.units_per_cleaning_cycle():,} units")
        print("="*64)

        print("\n[Phase 1] Preparing SKU table...")
        df_valid, df_all, all_machines = self._prepare_skus(
            df_demand, df_cycles, df_allow, df_gt, mould_tracker, df_running
        )

        print("\n[Phase 2] Building continuity blocks...")
        continuity_rows, locked_mins, demand_remainder = self._build_continuity(
            df_running, df_valid, df_gt, plan_start
        )

        # Reduce LP demand by what continuity already covers
        # SKUs fully handled in continuity are excluded from LP entirely
        df_lp = df_valid.copy()
        for sku, remainder in demand_remainder.items():
            mask = df_lp["SKUCode"] == sku
            if remainder == 0:
                # Continuity covers all demand — remove from LP
                df_lp = df_lp[~mask]
            else:
                # Continuity covers part — LP only needs the remainder
                df_lp.loc[mask, "Demand"] = remainder
                df_lp.loc[mask, "Demand_Mins"] = (
                    math.ceil(remainder / Config.CAVITIES_PER_MOULD)
                    * df_lp.loc[mask, "CycleTime_min"].values[0]
                )
        print(f"  [LP Input] {len(df_lp)} SKUs | "
              f"Remaining demand: {df_lp['Demand'].sum():,.0f} units")

        print("\n[Phase 3] Solving LP...")
        solver   = LP_Solver()
        x, meta  = solver.solve(df_lp, all_machines, mould_tracker, locked_mins)

        print("\n[Phase 4] Rounding to integer cycles...")
        rounder  = Rounder()
        df_mach, machine_sku_order = rounder.round(x, meta, df_lp, locked_mins)

        print("\n[Phase 5] Building shift-wise schedule...")
        builder  = ScheduleBuilder(plan_start)
        df_shift = builder.build(df_mach, machine_sku_order, df_gt, continuity_rows)
        df_shift.to_excel("df_shiftv1.xlsx", index=False)

        df_summary = self._build_summary(df_all, df_mach, df_shift)
        df_util    = self._build_util(df_mach, df_shift, all_machines)

        self._print_results(df_summary, df_util, df_shift)


        return {
            "machine_schedule":   df_mach,
            "shift_schedule":     df_shift,
            "demand_fulfillment": df_summary,
            "machine_utilization":df_util,
            "mould_tracker":      mould_tracker.summary,
        }


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER  (v2 — adds Changeover / Cleaning sheet)
# ══════════════════════════════════════════════════════════════════════════════
class ExcelExporter:
    _C = {"navy":"1F3864","blue":"2E75B6","teal":"1F6B75",
          "green":"C6EFCE","amber":"FFEB9C","red":"FFC7CE",
          "grey":"F2F2F2","white":"FFFFFF","lgrey":"E8E8E8","orange":"F4B942"}

    def __init__(self, path: str): self.path = path

    def F(self, c): return PatternFill("solid", fgColor=self._C.get(c, c))
    def _b(self):
        s=Side(style="thin",color="CCCCCC")
        return Border(left=s,right=s,top=s,bottom=s)
    def _hf(self): return Font(bold=True,name="Arial",size=10,color="FFFFFF")
    def _bf(self, bold=False): return Font(bold=bold,name="Arial",size=9)

    def _cell(self, ws, r, c, v, fmt=None, fc="white", bold=False, aln="center"):
        cell=ws.cell(r,c,v); cell.font=self._bf(bold); cell.fill=self.F(fc)
        cell.border=self._b()
        cell.alignment=Alignment(horizontal=aln,vertical="center",wrap_text=True)
        if fmt: cell.number_format=fmt

    def _hdr(self, ws, r, c, v, fc="navy"):
        cell=ws.cell(r,c,v); cell.font=self._hf(); cell.fill=self.F(fc)
        cell.border=self._b()
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

    def _title(self, ws, text, sub, n):
        ws.insert_rows(1); ws.insert_rows(1)
        cl=get_column_letter(n)
        ws.merge_cells(f"A1:{cl}1"); ws["A1"]=text
        ws["A1"].font=Font(bold=True,name="Arial",size=13,color="FFFFFF")
        ws["A1"].fill=self.F("navy"); ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height=26
        ws.merge_cells(f"A2:{cl}2"); ws["A2"]=sub
        ws["A2"].font=Font(italic=True,name="Arial",size=9,color="FFFFFF")
        ws["A2"].fill=self.F("teal"); ws["A2"].alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[2].height=16

    def _hdr_row(self, ws, row, n):
        for c in range(1,n+1):
            cell=ws.cell(row,c)
            cell.font=self._hf(); cell.fill=self.F("navy"); cell.border=self._b()
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.row_dimensions[row].height=30

    STATUS_FC = {"FULLY MET":"green","PARTIAL":"amber",
                 "UNMET":"red","UNSCHEDULABLE":"lgrey"}

    def export(self, results: dict):
        df_mach  = results["machine_schedule"]
        df_shift = results["shift_schedule"]
        df_sum   = results["demand_fulfillment"]
        df_util  = results["machine_utilization"]
        df_mould = results["mould_tracker"]

        td   = int(df_sum["Demand"].sum())
        tp   = int(df_sum["Planned_Units"].sum())
        tg   = int(df_sum["Gap"].sum())
        pct  = round(tp/td*100,1) if td else 0
        avg  = round(df_util["Utilization_Pct"].mean(),1)
        co_n = (df_shift["SKUCode"]=="CHANGEOVER").sum() if not df_shift.empty else 0
        cl_n = (df_shift["SKUCode"]=="MOULD_CLEAN").sum() if not df_shift.empty else 0
        kpi  = (f"Demand: {td:,}  |  Planned: {tp:,}  |  Gap: {tg:,}  |  "
                f"Fulfillment: {pct}%  |  Avg Util: {avg}%  |  "
                f"Changeovers: {co_n}  |  Mould Cleans: {cl_n}")

        with pd.ExcelWriter(self.path, engine="openpyxl") as writer:

            # ── Demand Fulfillment ────────────────────────────────────────────
            cols1=["SKUCode","Priority","Demand","GT_Inventory","Planned_Units",
                   "Gap","Fulfillment_Pct","Status","CycleTime_min",
                   "Eligible_Machines","Presses_Needed","Skip_Reason"]
            df_sum[cols1].to_excel(writer,sheet_name="Demand Fulfillment",index=False)
            ws=writer.book["Demand Fulfillment"]
            self._title(ws,"PCR CURING v2 — DEMAND FULFILLMENT",kpi,len(cols1))
            self._hdr_row(ws,3,len(cols1))
            for ci,w in enumerate([26,10,13,12,13,12,10,14,12,16,13,22],1):
                ws.column_dimensions[get_column_letter(ci)].width=w
            for ri in range(4,ws.max_row+1):
                st=str(ws.cell(ri,8).value)
                sf=self.STATUS_FC.get(st,"white"); bf="grey" if ri%2==0 else "white"
                for ci in range(1,len(cols1)+1):
                    self._cell(ws,ri,ci,ws.cell(ri,ci).value,
                               fc=sf if ci in(7,8) else bf,
                               bold=(ci==5),aln="left" if ci==1 else "center")
                fp=ws.cell(ri,7)
                if isinstance(fp.value,(int,float)): fp.value=fp.value/100
                fp.number_format="0.0%"
            tr=ws.max_row+1
            for ci in range(1,len(cols1)+1):
                c=ws.cell(tr,ci); c.fill=self.F("navy"); c.font=self._hf()
                c.border=self._b(); c.alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(tr,1).value="TOTAL"
            for ci,v,fmt in [(3,td,"#,##0"),(5,tp,"#,##0"),(6,tg,"#,##0"),(7,pct/100,"0.0%")]:
                ws.cell(tr,ci).value=v; ws.cell(tr,ci).number_format=fmt

            # ── Machine Schedule ──────────────────────────────────────────────
            cols2=["Machine","SKUCode","Priority","CycleTime_min",
                   "Cycles","Units_Planned","Mins_Used","Days_Used"]
            df_mach.sort_values(["Machine","SKUCode"]).to_excel(
                writer,sheet_name="Machine Schedule",index=False)
            ws2=writer.book["Machine Schedule"]
            self._title(ws2,"MACHINE-WISE SCHEDULE — PCR v2 (LP + CO + Cleaning)",kpi,len(cols2))
            self._hdr_row(ws2,3,len(cols2))
            for ci,w in enumerate([12,26,10,14,12,14,16,14],1):
                ws2.column_dimensions[get_column_letter(ci)].width=w
            prev=None
            for ri in range(4,ws2.max_row+1):
                m=ws2.cell(ri,1).value; bf="lgrey" if m!=prev else ("grey" if ri%2==0 else "white")
                prev=m
                for ci in range(1,len(cols2)+1):
                    self._cell(ws2,ri,ci,ws2.cell(ri,ci).value,
                               fc=bf,bold=(ci in(1,6)),aln="left" if ci==2 else "center")

            # ── Shift Schedule (with CHANGEOVER / MOULD_CLEAN rows) ───────────
            cols3=["Date","Shift","Machine","SKUCode","StartTime","EndTime",
                   "Qty","CycleTime_min","GT_Inventory","Remarks"]
            df_shift[cols3].to_excel(writer,sheet_name="Shift Schedule",index=False)
            ws3=writer.book["Shift Schedule"]
            self._title(ws3,"SHIFT-WISE SCHEDULE — PCR v2",kpi,len(cols3))
            self._hdr_row(ws3,3,len(cols3))
            for ci,w in enumerate([12,8,12,26,18,18,10,12,12,26],1):
                ws3.column_dimensions[get_column_letter(ci)].width=w
            ROW_FC={"CHANGEOVER":"orange","MOULD_CLEAN":"amber",
                    "A":"E8F4F8","B":"FFF8E8","C":"F0F0F0"}
            for ri in range(4,ws3.max_row+1):
                sku=str(ws3.cell(ri,4).value); shift=str(ws3.cell(ri,2).value)
                fc=ROW_FC.get(sku,ROW_FC.get(shift,"white"))
                for ci in range(1,len(cols3)+1):
                    self._cell(ws3,ri,ci,ws3.cell(ri,ci).value,
                               fc=fc,bold=(sku in("CHANGEOVER","MOULD_CLEAN")),
                               aln="left" if ci==4 else "center")

            # ── Machine Utilization ───────────────────────────────────────────
            cols4=["Machine","Available_Mins","Used_Mins","Idle_Mins",
                   "Utilization_Pct","SKUs_Count","Total_Cycles","Total_Units"]
            df_util.to_excel(writer,sheet_name="Machine Utilization",index=False)
            ws4=writer.book["Machine Utilization"]
            idle_c=int((df_util["Utilization_Pct"]==0).sum())
            high_c=int((df_util["Utilization_Pct"]>=90).sum())
            self._title(ws4,"PRESS UTILIZATION — PCR v2",
                        f"Avg: {avg}% | High(≥90%): {high_c} | Idle: {idle_c} | Total: {len(df_util)}",
                        len(cols4))
            self._hdr_row(ws4,3,len(cols4))
            for ci,w in enumerate([12,15,14,14,14,12,14,14],1):
                ws4.column_dimensions[get_column_letter(ci)].width=w
            for ri in range(4,ws4.max_row+1):
                u=ws4.cell(ri,5).value or 0
                uf="green" if u>=90 else ("amber" if u>=60 else "red")
                bf="grey" if ri%2==0 else "white"
                for ci in range(1,len(cols4)+1):
                    self._cell(ws4,ri,ci,ws4.cell(ri,ci).value,
                               fc=uf if ci==5 else bf,bold=(ci in(1,5)))
                fp=ws4.cell(ri,5)
                if isinstance(fp.value,(int,float)): fp.value=fp.value/100
                fp.number_format="0.0%"

            # ── Mould Tracker ─────────────────────────────────────────────────
            df_mould.to_excel(writer,sheet_name="Mould Tracker",index=False)
            ws5=writer.book["Mould Tracker"]
            self._title(ws5,"MOULD AVAILABILITY TRACKER",
                        f"Total moulds: {len(df_mould)} | "
                        f"Free: {(df_mould['Assigned_Machine']=='FREE').sum()} | "
                        f"Assigned: {(df_mould['Assigned_Machine']!='FREE').sum()}",
                        len(df_mould.columns))
            self._hdr_row(ws5,3,len(df_mould.columns))
            for ci,w in enumerate([22,30,14,16],1):
                ws5.column_dimensions[get_column_letter(ci)].width=w
            for ri in range(4,ws5.max_row+1):
                assigned=str(ws5.cell(ri,4).value)
                bf="C6EFCE" if assigned=="FREE" else ("FFEB9C" if assigned!="FREE" else "white")
                for ci in range(1,len(df_mould.columns)+1):
                    self._cell(ws5,ri,ci,ws5.cell(ri,ci).value,fc=bf)

        print(f"\n  [Export] Saved → {self.path}")



def con_split_into_shifts(df):
    rows = []

    for _, r in df.iterrows():
        start = r['StartTime']
        end   = r['EndTime']
        total_qty = r['Qty']
        total_minutes = (end - start).total_seconds() / 60

        current = start

        while current < end:
            shift, shift_end = _get_shift_fn(current)

            slice_end = min(shift_end, end)
            slice_minutes = (slice_end - current).total_seconds() / 60

            # proportional qty allocation
            qty = (slice_minutes / total_minutes) * total_qty if total_minutes > 0 else 0

            new_row = r.copy()
            new_row['StartTime'] = current
            new_row['EndTime']   = slice_end
            new_row['Shift']     = shift
            new_row['Qty']       = round(qty)

            rows.append(new_row)

            current = slice_end

    return pd.DataFrame(rows)
# ══════════════════════════════════════════════════════════════════════════════
# PATCH: make _get_shift a static-callable helper
# ══════════════════════════════════════════════════════════════════════════════
def _get_shift_fn(dt: datetime) -> tuple[str, datetime]:
    h    = dt.hour
    base = dt.replace(hour=0, minute=0, second=0, microsecond=0)
    sh   = Config.SHIFT_START_HOUR
    if sh <= h < sh + 8:
        return "A", base + timedelta(hours=sh + 8)
    elif sh + 8 <= h < sh + 16:
        return "B", base + timedelta(hours=sh + 16)
    else:
        if h >= sh + 16:
            return "C", base + timedelta(days=1, hours=sh)
        else:
            return "C", base + timedelta(hours=sh)
# ══════════════════════════════════════════════════════════════════════════════
# ENTRY POINTS
# ══════════════════════════════════════════════════════════════════════════════
def run_from_excel(
    demand_path:  str = "Demand_for_Curing_Schedule3_pcr.xlsx",
    cycles_path:  str = "Master_Curing_Design_CycleTime_pcr.xlsx",
    allow_path:   str = "curing_pcr_machine_allowable.xlsx",
    gt_path:      str = "GT_Inventory_pcr.xlsx",
    mould_path:   str = "Master_Mapping_Mould_SKU.xlsx",
    running_path: str = None,
    plan_start:   datetime = None,
    output_path:  str = "PCR_Curing_LP_v2_Schedule.xlsx",
) -> dict:
    print("\n[Phase 0] ETL from Excel files...")
    df_demand  = ETL.load_demand_from_excel(demand_path)
    df_cycles  = ETL.load_cycle_times_from_excel(cycles_path)
    df_allow   = ETL.load_machine_allowable_from_excel(allow_path)
    df_gt      = ETL.load_gt_inventory_from_excel(gt_path)
    df_running = ETL.load_running_moulds_from_excel(running_path) if running_path else None

    tracker = MouldTracker()
    tracker.load_from_excel(mould_path, running_path)

    scheduler = JK_LP_Curing_Scheduler_v2()
    results   = scheduler.run(df_demand, df_cycles, df_allow, df_gt,
                              tracker, df_running, plan_start)
    ExcelExporter(output_path).export(results)
    return results


def run_from_database(
    demand_csv:   str,
    plan_start:   datetime = None,
    tyre_type:    str = Config.TYRE_TYPE,
    output_path:  str = Config.OUTPUT_FILE,
) -> dict:
    if create_engine is None:
        raise ImportError("sqlalchemy not installed. Use run_from_excel() instead.")
    engine = create_engine(
        f"mysql+pymysql://{Config.DB_USER}:{Config.DB_PASSWORD}"
        f"@{Config.DB_SERVER}/{Config.DB_NAME}"
    )
    etl = ETL(engine, tyre_type)
    print("\n[Phase 0] ETL from database...")
    df_demand  = etl.load_demand(demand_csv)
    df_cycles  = etl.load_cycle_times()
    df_allow   = etl.load_machine_allowable()
    df_gt      = etl.load_gt_inventory()
    df_running = etl.load_running_moulds()
    df_mould_m = etl.load_mould_master()

    tracker = MouldTracker()
    tracker.load_from_df(df_mould_m, df_running)

    scheduler = JK_LP_Curing_Scheduler_v2()
    results   = scheduler.run(df_demand, df_cycles, df_allow, df_gt,
                              tracker, df_running, plan_start)
    ExcelExporter(output_path).export(results)
    return results


# ------- DB Execution ----------------------------------------------------------

if __name__ == "__main__":
    results = run_from_database(
        demand_csv = "Feb_CTP_PCR_Requirement.csv",
        plan_start = Config.PLAN_DATE#datetime(2026, 2, 1, 7, 0, 0)
    )
# ------- Excel Execution -------------------------------------------------------
# ══════════════════════════════════════════════════════════════════════════════
'''if __name__ == "__main__":
    results = run_from_excel(
        demand_path  = "Demand_for_Curing_Schedule3_pcr.xlsx",
        cycles_path  = "Master_Curing_Design_CycleTime_pcr.xlsx",
        allow_path   = "curing_pcr_machine_allowable.xlsx",
        gt_path      = "GT_Inventory_pcr.xlsx",
        mould_path   = "Master_Mapping_Mould_SKU.xlsx",
        running_path = None,   # set to "Curing_Current_Running_moulds_pcr.xlsx" for continuity
        plan_start   = datetime(2026, 4, 1, 7, 0, 0),
        output_path  = "PCR_Curing_LP_v2_Schedule.xlsx",
    )'''
