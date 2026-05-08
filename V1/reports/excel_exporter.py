"""
ExcelExporter — writes the 5-sheet PlanSchedule workbook (Demand Fulfillment,
Machine Schedule, Shift Schedule, Machine Utilization, Mould Tracker).
"""

import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


# ══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORTER  (v2 — adds Changeover / Cleaning sheet)
# ══════════════════════════════════════════════════════════════════════════════
class ExcelExporter:
    _C = {"navy":"1F3864","blue":"2E75B6","teal":"1F6B75",
          "green":"C6EFCE","amber":"FFEB9C","red":"FFC7CE",
          "grey":"F2F2F2","white":"FFFFFF","lgrey":"E8E8E8","orange":"F4B942"}

    # Algorithm-specific title fragments. The strings below preserve the
    # exact wording used by the original LP/MILP/CP-SAT monoliths so the
    # generated workbook is byte-equivalent to the legacy outputs.
    #
    #   variant   -> "v2"                       (LP) | "MILP v1" | "CP-SAT v1"
    #   ms_suffix -> " (LP + CO + Cleaning)"    (LP) | " (CO + Cleaning)" (others)
    _TITLE_VARIANTS = {
        "LP":     {"variant": "v2",         "ms_suffix": " (LP + CO + Cleaning)"},
        "MILP":   {"variant": "MILP v1",    "ms_suffix": " (CO + Cleaning)"},
        "CP-SAT": {"variant": "CP-SAT v1",  "ms_suffix": " (CO + Cleaning)"},
    }

    def __init__(self, path: str, algo_label: str = "LP"):
        self.path = path
        cfg = self._TITLE_VARIANTS.get(algo_label, self._TITLE_VARIANTS["LP"])
        v = cfg["variant"]
        suf = cfg["ms_suffix"]
        if algo_label == "LP":
            self._title_demand = f"PCR CURING {v} — DEMAND FULFILLMENT"
        else:
            self._title_demand = f"PCR CURING {v} — DEMAND FULFILLMENT"
        self._title_machine = f"MACHINE-WISE SCHEDULE — PCR {v}{suf}"
        self._title_shift   = f"SHIFT-WISE SCHEDULE — PCR {v}"
        self._title_util    = f"PRESS UTILIZATION — PCR {v}"

    def F(self, c): return PatternFill("solid", fgColor=self._C.get(c, c))
    def _b(self):
        s=Side(style="thin",color="CCCCCC")
        return Border(left=s,right=s,top=s,bottom=s)
    def _hf(self): return Font(bold=True,name="Arial",size=10,color="FFFFFF")
    def _bf(self, bold=False): return Font(bold=bold,name="Arial",size=9)

    def _cell(self, ws, r, c, v, fmt=None, fc="white", bold=False, aln="center"):
        cell=ws.cell(r,c,v); cell.font=self._bf(bold); cell.fill=self.F(fc)
        cell.border=self._b()
        cell.alignment=Alignment(horizontal=aln,vertical="center",wrap_text=True)
        if fmt: cell.number_format=fmt

    def _hdr(self, ws, r, c, v, fc="navy"):
        cell=ws.cell(r,c,v); cell.font=self._hf(); cell.fill=self.F(fc)
        cell.border=self._b()
        cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)

    def _title(self, ws, text, sub, n):
        ws.insert_rows(1); ws.insert_rows(1)
        cl=get_column_letter(n)
        ws.merge_cells(f"A1:{cl}1"); ws["A1"]=text
        ws["A1"].font=Font(bold=True,name="Arial",size=13,color="FFFFFF")
        ws["A1"].fill=self.F("navy"); ws["A1"].alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[1].height=26
        ws.merge_cells(f"A2:{cl}2"); ws["A2"]=sub
        ws["A2"].font=Font(italic=True,name="Arial",size=9,color="FFFFFF")
        ws["A2"].fill=self.F("teal"); ws["A2"].alignment=Alignment(horizontal="center",vertical="center")
        ws.row_dimensions[2].height=16

    def _hdr_row(self, ws, row, n):
        for c in range(1,n+1):
            cell=ws.cell(row,c)
            cell.font=self._hf(); cell.fill=self.F("navy"); cell.border=self._b()
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True)
        ws.row_dimensions[row].height=30

    STATUS_FC = {"FULLY MET":"green","PARTIAL":"amber",
                 "UNMET":"red","UNSCHEDULABLE":"lgrey"}

    def export(self, results: dict):
        df_mach  = results["machine_schedule"]
        df_shift = results["shift_schedule"]
        df_sum   = results["demand_fulfillment"]
        df_util  = results["machine_utilization"]
        df_mould = results["mould_tracker"]

        td   = int(df_sum["Demand"].sum())
        tp   = int(df_sum["Planned_Units"].sum())
        tg   = int(df_sum["Gap"].sum())
        pct  = round(tp/td*100,1) if td else 0
        avg  = round(df_util["Utilization_Pct"].mean(),1)
        co_n = (df_shift["SKUCode"]=="CHANGEOVER").sum() if not df_shift.empty else 0
        cl_n = (df_shift["SKUCode"]=="MOULD_CLEAN").sum() if not df_shift.empty else 0
        kpi  = (f"Demand: {td:,}  |  Planned: {tp:,}  |  Gap: {tg:,}  |  "
                f"Fulfillment: {pct}%  |  Avg Util: {avg}%  |  "
                f"Changeovers: {co_n}  |  Mould Cleans: {cl_n}")

        with pd.ExcelWriter(self.path, engine="openpyxl") as writer:

            # ── Demand Fulfillment ────────────────────────────────────────────
            cols1=["SKUCode","Priority","Demand","GT_Inventory","Planned_Units",
                   "Gap","Fulfillment_Pct","Status","CycleTime_min",
                   "Eligible_Machines","Presses_Needed","Skip_Reason"]
            df_sum[cols1].to_excel(writer,sheet_name="Demand Fulfillment",index=False)
            ws=writer.book["Demand Fulfillment"]
            self._title(ws,self._title_demand,kpi,len(cols1))
            self._hdr_row(ws,3,len(cols1))
            for ci,w in enumerate([26,10,13,12,13,12,10,14,12,16,13,22],1):
                ws.column_dimensions[get_column_letter(ci)].width=w
            for ri in range(4,ws.max_row+1):
                st=str(ws.cell(ri,8).value)
                sf=self.STATUS_FC.get(st,"white"); bf="grey" if ri%2==0 else "white"
                for ci in range(1,len(cols1)+1):
                    self._cell(ws,ri,ci,ws.cell(ri,ci).value,
                               fc=sf if ci in(7,8) else bf,
                               bold=(ci==5),aln="left" if ci==1 else "center")
                fp=ws.cell(ri,7)
                if isinstance(fp.value,(int,float)): fp.value=fp.value/100
                fp.number_format="0.0%"
            tr=ws.max_row+1
            for ci in range(1,len(cols1)+1):
                c=ws.cell(tr,ci); c.fill=self.F("navy"); c.font=self._hf()
                c.border=self._b(); c.alignment=Alignment(horizontal="center",vertical="center")
            ws.cell(tr,1).value="TOTAL"
            for ci,v,fmt in [(3,td,"#,##0"),(5,tp,"#,##0"),(6,tg,"#,##0"),(7,pct/100,"0.0%")]:
                ws.cell(tr,ci).value=v; ws.cell(tr,ci).number_format=fmt

            # ── Machine Schedule ──────────────────────────────────────────────
            cols2=["Machine","SKUCode","Priority","CycleTime_min",
                   "Cycles","Units_Planned","Mins_Used","Days_Used"]
            df_mach.sort_values(["Machine","SKUCode"]).to_excel(
                writer,sheet_name="Machine Schedule",index=False)
            ws2=writer.book["Machine Schedule"]
            self._title(ws2,self._title_machine,kpi,len(cols2))
            self._hdr_row(ws2,3,len(cols2))
            for ci,w in enumerate([12,26,10,14,12,14,16,14],1):
                ws2.column_dimensions[get_column_letter(ci)].width=w
            prev=None
            for ri in range(4,ws2.max_row+1):
                m=ws2.cell(ri,1).value; bf="lgrey" if m!=prev else ("grey" if ri%2==0 else "white")
                prev=m
                for ci in range(1,len(cols2)+1):
                    self._cell(ws2,ri,ci,ws2.cell(ri,ci).value,
                               fc=bf,bold=(ci in(1,6)),aln="left" if ci==2 else "center")

            # ── Shift Schedule (with CHANGEOVER / MOULD_CLEAN rows) ───────────
            cols3=["Date","Shift","Machine","SKUCode","StartTime","EndTime",
                   "Qty","CycleTime_min","GT_Inventory","Remarks"]
            df_shift[cols3].to_excel(writer,sheet_name="Shift Schedule",index=False)
            ws3=writer.book["Shift Schedule"]
            self._title(ws3,self._title_shift,kpi,len(cols3))
            self._hdr_row(ws3,3,len(cols3))
            for ci,w in enumerate([12,8,12,26,18,18,10,12,12,26],1):
                ws3.column_dimensions[get_column_letter(ci)].width=w
            ROW_FC={"CHANGEOVER":"orange","MOULD_CLEAN":"amber",
                    "A":"E8F4F8","B":"FFF8E8","C":"F0F0F0"}
            for ri in range(4,ws3.max_row+1):
                sku=str(ws3.cell(ri,4).value); shift=str(ws3.cell(ri,2).value)
                fc=ROW_FC.get(sku,ROW_FC.get(shift,"white"))
                for ci in range(1,len(cols3)+1):
                    self._cell(ws3,ri,ci,ws3.cell(ri,ci).value,
                               fc=fc,bold=(sku in("CHANGEOVER","MOULD_CLEAN")),
                               aln="left" if ci==4 else "center")

            # ── Machine Utilization ───────────────────────────────────────────
            cols4=["Machine","Available_Mins","Used_Mins","Idle_Mins",
                   "Utilization_Pct","SKUs_Count","Total_Cycles","Total_Units"]
            df_util.to_excel(writer,sheet_name="Machine Utilization",index=False)
            ws4=writer.book["Machine Utilization"]
            idle_c=int((df_util["Utilization_Pct"]==0).sum())
            high_c=int((df_util["Utilization_Pct"]>=90).sum())
            self._title(ws4,self._title_util,
                        f"Avg: {avg}% | High(≥90%): {high_c} | Idle: {idle_c} | Total: {len(df_util)}",
                        len(cols4))
            self._hdr_row(ws4,3,len(cols4))
            for ci,w in enumerate([12,15,14,14,14,12,14,14],1):
                ws4.column_dimensions[get_column_letter(ci)].width=w
            for ri in range(4,ws4.max_row+1):
                u=ws4.cell(ri,5).value or 0
                uf="green" if u>=90 else ("amber" if u>=60 else "red")
                bf="grey" if ri%2==0 else "white"
                for ci in range(1,len(cols4)+1):
                    self._cell(ws4,ri,ci,ws4.cell(ri,ci).value,
                               fc=uf if ci==5 else bf,bold=(ci in(1,5)))
                fp=ws4.cell(ri,5)
                if isinstance(fp.value,(int,float)): fp.value=fp.value/100
                fp.number_format="0.0%"

            # ── Mould Tracker ─────────────────────────────────────────────────
            df_mould.to_excel(writer,sheet_name="Mould Tracker",index=False)
            ws5=writer.book["Mould Tracker"]
            self._title(ws5,"MOULD AVAILABILITY TRACKER",
                        f"Total moulds: {len(df_mould)} | "
                        f"Free: {(df_mould['Assigned_Machine']=='FREE').sum()} | "
                        f"Assigned: {(df_mould['Assigned_Machine']!='FREE').sum()}",
                        len(df_mould.columns))
            self._hdr_row(ws5,3,len(df_mould.columns))
            for ci,w in enumerate([22,30,14,16],1):
                ws5.column_dimensions[get_column_letter(ci)].width=w
            for ri in range(4,ws5.max_row+1):
                assigned=str(ws5.cell(ri,4).value)
                bf="C6EFCE" if assigned=="FREE" else ("FFEB9C" if assigned!="FREE" else "white")
                for ci in range(1,len(df_mould.columns)+1):
                    self._cell(ws5,ri,ci,ws5.cell(ri,ci).value,fc=bf)

        print(f"\n  [Export] Saved → {self.path}")
